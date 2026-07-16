from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import streamlit as st
import pdfplumber
import fitz
import json
import re
import anthropic
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

st.set_page_config(page_title="BRSR Gap Analysis Tool", layout="wide")
st.markdown("""
<style>
    .block-container {
        padding-top: 2.5rem;
        padding-bottom: 3rem;
    }
    h1 {
        font-weight: 800 !important;
        letter-spacing: -0.5px;
        background: linear-gradient(90deg, #2ecc71, #3498db);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    div[data-testid="stMetric"] {
        background-color: #1a1f2b !important;
        border: 1px solid #2a2f3b;
        border-radius: 12px;
        padding: 20px 15px;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    div[data-testid="stMetric"] label,
    div[data-testid="stMetric"] div {
        color: #fafafa !important;
    }
    div[data-testid="stMetric"]:hover {
        transform: translateY(-4px);
        box-shadow: 0 8px 20px rgba(46, 204, 113, 0.15);
    }
    h2, h3 {
        border-bottom: 2px solid #2ecc71;
        padding-bottom: 8px;
        display: inline-block;
    }
    div[data-testid="stFileUploader"] {
        border-radius: 12px;
    }
    div[data-testid="stTable"] {
        border-radius: 10px;
        overflow: hidden;
    }
    section[data-testid="stSidebar"] {
        background-color: #14181f !important;
        border-right: 1px solid #2a2f3b;
    }
    section[data-testid="stSidebar"] * {
        color: #fafafa !important;
    }
    html {
        scroll-behavior: smooth;
    }
</style>
""", unsafe_allow_html=True)

with st.sidebar:
    st.header("ℹ️ About this tool")
    st.write(
        "This tool scans any company's BRSR (Business Responsibility and "
        "Sustainability Report) filing and checks it against SEBI's "
        "**BRSR Core** — the 9 ESG attributes requiring third-party assurance."
    )
    st.divider()
    st.subheader("How to use")
    st.markdown(
        "1. Upload a company's Annual Report or standalone BRSR PDF\n"
        "2. The tool locates the Principle-wise disclosure section\n"
        "3. Each disclosure is checked against BRSR Core requirements\n"
        "4. View results as charts, tables, or a downloadable report"
    )
    st.divider()
    st.subheader("🤖 AI Narratives (optional)")
    _existing_key = st.secrets.get("ANTHROPIC_API_KEY", None) if hasattr(st, "secrets") else None
    if _existing_key:
        st.caption("Anthropic API key loaded from app secrets.")
    else:
        st.text_input(
            "Anthropic API key",
            type="password",
            key="anthropic_api_key",
            help="Needed only for AI-generated compliance gap narratives. Not stored anywhere."
        )
    st.divider()
    st.markdown(
        "Built by **Karan Thakur** · "
        "[GitHub](https://github.com/karan02566-prog) · "
        "[LinkedIn](https://www.linkedin.com/in/karan-thakur-3486b538a/)"
    )

st.markdown("""
<style>
    .hero-box {
        background: linear-gradient(135deg, #0d1f14, #0a1a25);
        border: 1px solid #2a3f2f;
        border-radius: 16px;
        padding: 32px 36px;
        margin-bottom: 24px;
    }
    .hero-badge {
        display: inline-block;
        background: rgba(46, 204, 113, 0.15);
        color: #2ecc71;
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 0.5px;
        padding: 5px 12px;
        border-radius: 20px;
        margin-bottom: 14px;
        text-transform: uppercase;
    }
    .hero-title {
        font-size: 34px;
        font-weight: 800;
        line-height: 1.25;
        margin: 0 0 12px 0;
        background: linear-gradient(90deg, #ffffff, #2ecc71);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .hero-sub {
        font-size: 15px;
        color: #c9d1d9;
        max-width: 720px;
        line-height: 1.5;
        margin: 0;
    }
</style>
<div class="hero-box">
    <span class="hero-badge">Built for auditors, analysts &amp; ESG reviewers</span>
    <p class="hero-title">Check any listed company's BRSR filing against SEBI's Core checklist — with the evidence to prove it.</p>
    <p class="hero-sub">
        Upload a company's Annual Report or standalone BRSR PDF. This tool locates the Principle-wise
        disclosure section, checks it against SEBI's 9 BRSR Core attributes requiring third-party assurance,
        and shows you the exact sentence behind every match — including catching disclosures that
        <i>mention</i> a metric but actually deny it (e.g. "no water recycling program"). No black-box scoring.
    </p>
</div>
""", unsafe_allow_html=True)

stat_col1, stat_col2, stat_col3 = st.columns(3)
stat_col1.metric("BRSR Core attributes checked", "9")
stat_col2.metric("Time to full analysis", "~30 sec")
stat_col3.metric("Negation-aware matching", "Yes")

with st.expander("📖 How the scoring works (methodology, for reviewers)"):
    st.markdown("""
**1. Section detection** — the tool scans the uploaded PDF for BRSR section markers
("Section C: Principle Wise Performance", "Principle 1:" through "Principle 9:") and
extracts the Principle-wise disclosure text using `pdfplumber`.

**2. Keyword matching per attribute** — each of the 9 BRSR Core attributes has a defined
set of keywords (e.g. *GHG Footprint* → `scope 1`, `scope 2`, `scope 3`, `ghg`, `emission intensity`).
The full reference checklist is in `brsr_core_reference.json`, mapped to its BRSR Principle.

**3. Negation check** — a raw keyword hit isn't enough. The tool splits the filing into
sentences and checks the ~8 words before each keyword match for negation language
("no", "not", "lack of", "excluding", "does not", etc.). A negated mention is
excluded from coverage and flagged separately as evidence, not counted as a false positive.

**4. Scoring** — coverage % = (keywords genuinely matched) ÷ (total keywords for that attribute).
≥50% → ✅ Covered · >0% → 🟡 Partially Covered · 0% → 🔴 Missing.

**5. Evidence** — every match (or negated non-match) links back to the actual sentence it
came from, shown in the Detailed Coverage section below, so you can verify the score
yourself rather than trusting a black box.

*Limitations: this is a rule-based keyword/context matcher, not a semantic AI reader — it
can still miss disclosures phrased with unlisted synonyms. Treat scores as a fast first-pass
gap check, not a substitute for a manual audit.*
""")

st.write("")

# --- Load the reference checklist (same one from Module 3) ---
with open("brsr_core_reference.json", "r") as f:
    REFERENCE = json.load(f)


# --- Reused functions from Module 1 ---
def find_brsr_section(doc, keywords=None):
    if keywords is None:
        keywords = [
            "Business Responsibility and Sustainability Report",
            "Section A: General Disclosures",
            "Section B: Management and Process",
            "Section C: Principle Wise Performance"
        ]
    hits = []
    for page_num in range(len(doc)):
        text = doc[page_num].get_text()
        for kw in keywords:
            if kw.lower() in text.lower():
                hits.append({"page": page_num, "keyword": kw})
    return hits


def extract_text_pdfplumber(pdf_file, start_page, end_page):
    extracted = {}
    with pdfplumber.open(pdf_file) as pdf:
        pages = pdf.pages[start_page:end_page]
        for i, page in enumerate(pages):
            text = page.extract_text() or ""
            extracted[start_page + i] = {"text": text}
    return extracted


def chunk_by_principle(extracted_pages):
    principle_markers = [f"Principle {i}:" for i in range(1, 10)]
    full_text = "\n".join([p["text"] for p in extracted_pages.values()])
    chunks = {}
    positions = []
    for marker in principle_markers:
        for match in re.finditer(re.escape(marker), full_text, re.IGNORECASE):
            positions.append((match.start(), marker))
    positions.sort()
    for idx, (pos, marker) in enumerate(positions):
        end = positions[idx + 1][0] if idx + 1 < len(positions) else len(full_text)
        if marker not in chunks:
            chunks[marker] = full_text[pos:end]
    return chunks


NEGATION_PATTERNS = [
    r"\bno\b", r"\bnot\b", r"\bnone\b", r"\bnil\b", r"\bwithout\b",
    r"\bn't\b", r"\bexcluding\b", r"\bexcludes\b", r"\babsence of\b",
    r"\black of\b", r"\bdoes not\b", r"\bdid not\b", r"\bhas not\b",
    r"\bhave not\b", r"\bhas no\b", r"\bhave no\b", r"\bnot applicable\b",
    r"\bnot yet\b", r"\bnot been\b", r"\bunable to\b", r"\bfailed to\b"
]
NEGATION_REGEX = re.compile("|".join(NEGATION_PATTERNS), re.IGNORECASE)


def split_sentences(text):
    """Rough sentence splitter — good enough for BRSR-style disclosure prose."""
    text = re.sub(r"\s+", " ", text)
    sentences = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", text)
    return [s.strip() for s in sentences if s.strip()]


def find_keyword_evidence(sentences, keyword):
    """Find sentences containing keyword; flag as negated if a negation word
    appears within ~8 words before the keyword in that sentence."""
    hits = []
    kw_lower = keyword.lower()
    for sentence in sentences:
        s_lower = sentence.lower()
        idx = s_lower.find(kw_lower)
        if idx == -1:
            continue
        window_start = max(0, idx - 60)  # ~8-10 words of lookback
        window = s_lower[window_start:idx]
        is_negated = bool(NEGATION_REGEX.search(window))
        hits.append({"sentence": sentence.strip(), "negated": is_negated})
    return hits


def run_gap_analysis(all_text):
    sentences = split_sentences(all_text)
    all_text_lower = all_text.lower()
    results = []
    for attribute, details in REFERENCE.items():
        matched_keywords = []
        missing_keywords = []
        evidence = {}
        for kw in details["keywords"]:
            if kw not in all_text_lower:
                missing_keywords.append(kw)
                continue
            hits = find_keyword_evidence(sentences, kw)
            positive_hits = [h for h in hits if not h["negated"]]
            if positive_hits:
                matched_keywords.append(kw)
                evidence[kw] = positive_hits[0]["sentence"]
            else:
                # keyword exists in text but every occurrence was negated
                missing_keywords.append(kw)
                if hits:
                    evidence[kw] = hits[0]["sentence"]  # keep as "negated evidence" for transparency
        coverage = len(matched_keywords) / len(details["keywords"])
        if coverage >= 0.5:
            status = "✅ Covered"
        elif coverage > 0:
            status = "🟡 Partially Covered"
        else:
            status = "🔴 Missing"
        results.append({
            "Attribute": attribute,
            "Principle": details["principle"],
            "Status": status,
            "Match %": round(coverage * 100, 1),
            "Matched Keywords": matched_keywords,
            "Missing Keywords": missing_keywords,
            "Evidence": evidence
        })
    return results


def get_anthropic_client():
    """Return an Anthropic client using an API key from secrets or sidebar input."""
    api_key = st.secrets.get("ANTHROPIC_API_KEY", None) if hasattr(st, "secrets") else None
    if not api_key:
        api_key = st.session_state.get("anthropic_api_key")
    if not api_key:
        return None
    return anthropic.Anthropic(api_key=api_key)


def generate_gap_narrative(client, attribute, principle, status, match_pct, matched_keywords, missing_keywords, company_name):
    """Call Claude to produce a specific, auditor-style narrative for one BRSR Core attribute gap."""
    prompt = f"""You are a BRSR (Business Responsibility and Sustainability Report) compliance analyst reviewing {company_name}'s disclosures against SEBI's BRSR Core requirements.

Attribute: {attribute}
Principle: {principle}
Current status: {status} ({match_pct}% keyword coverage)
Disclosure elements found: {", ".join(matched_keywords) if matched_keywords else "none"}
Disclosure elements NOT found: {", ".join(missing_keywords) if missing_keywords else "none"}

Write a short compliance gap narrative in exactly 3 parts, each 1-2 sentences, plain text (no markdown headers, no bullet symbols):
1. What is missing or incomplete in the disclosure for this attribute.
2. Why this matters for SEBI BRSR Core reasonable assurance requirements.
3. A specific, actionable recommendation for what the company should disclose next reporting cycle to close the gap.

Keep the entire response under 120 words. Be specific to the attribute, not generic."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=350,
            messages=[{"role": "user", "content": prompt}]
        )
        return response.content[0].text.strip()
    except Exception as e:
        return f"⚠️ Could not generate narrative: {e}"


def generate_pdf_report(results, overall_score, covered_count, partial_count, missing_count, company_name="Company", narratives=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "TitleStyle", parent=styles["Heading1"], textColor=colors.HexColor("#2ecc71")
    )
    elements.append(Paragraph("BRSR Core Gap Analysis Report", title_style))
    elements.append(Paragraph(f"Company: {company_name}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    elements.append(Paragraph(f"Overall BRSR Core Score: {overall_score}%", styles["Heading2"]))
    elements.append(Paragraph(
        f"Covered: {covered_count}   |   Partial: {partial_count}   |   Missing: {missing_count}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Detailed Attribute Coverage", styles["Heading2"]))
    elements.append(Spacer(1, 10))

    table_data = [["Attribute", "Principle", "Status", "Match %"]]
    for r in results:
        status_clean = r["Status"].replace("✅ ", "").replace("🟡 ", "").replace("🔴 ", "")
        table_data.append([r["Attribute"], r["Principle"], status_clean, f"{r['Match %']}%"])

    table = Table(table_data, colWidths=[2.3 * inch, 1 * inch, 1.2 * inch, 1 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2ecc71")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    if narratives:
        elements.append(Spacer(1, 24))
        elements.append(Paragraph("AI-Generated Compliance Gap Narratives", styles["Heading2"]))
        elements.append(Spacer(1, 10))
        narrative_style = ParagraphStyle(
            "NarrativeStyle", parent=styles["Normal"], spaceAfter=14
        )
        heading_style = ParagraphStyle(
            "NarrativeHeading", parent=styles["Heading3"], textColor=colors.HexColor("#2ecc71")
        )
        for r in results:
            text = narratives.get(r["Attribute"])
            if text:
                elements.append(Paragraph(f"{r['Attribute']} ({r['Principle']})", heading_style))
                elements.append(Paragraph(text.replace("\n", "<br/>"), narrative_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer
def generate_excel_report(results, overall_score, covered_count, partial_count, missing_count, company_name="Company"):
    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    summary_sheet["A1"] = "BRSR Core Gap Analysis Report"
    summary_sheet["A1"].font = Font(size=16, bold=True, color="2ECC71")
    summary_sheet["A2"] = f"Company: {company_name}"
    summary_sheet["A2"].font = Font(size=11, italic=True)

    summary_sheet["A4"] = "Overall BRSR Core Score"
    summary_sheet["B4"] = f"{overall_score}%"
    summary_sheet["A5"] = "Covered"
    summary_sheet["B5"] = covered_count
    summary_sheet["A6"] = "Partial"
    summary_sheet["B6"] = partial_count
    summary_sheet["A7"] = "Missing"
    summary_sheet["B7"] = missing_count

    for row in range(4, 8):
        summary_sheet[f"A{row}"].font = Font(bold=True)
        # ---------- Pie chart: Covered / Partial / Missing ----------
    pie = PieChart()
    pie.title = "Status Breakdown"

    labels = Reference(summary_sheet, min_col=1, min_row=5, max_row=7)   # A5:A7 -> "Covered","Partial","Missing"
    data = Reference(summary_sheet, min_col=2, min_row=5, max_row=7)     # B5:B7 -> the actual counts

    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12

    summary_sheet.add_chart(pie, "D4")

    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 15

    # ---------- Sheet 2: Detailed Gaps ----------
    detail_sheet = wb.create_sheet("Detailed Gaps")

    headers = ["Attribute", "Principle", "Status", "Match %"]
    detail_sheet.append(headers)

    # Style the header row
    header_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = detail_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Color fills for status
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Write each row of results
    for r in results:
        status_clean = r["Status"].replace("✅ ", "").replace("🟡 ", "").replace("🔴 ", "")
        row_data = [r["Attribute"], r["Principle"], status_clean, r["Match %"]]
        detail_sheet.append(row_data)

        current_row = detail_sheet.max_row
        status_cell = detail_sheet.cell(row=current_row, column=3)

        if "Covered" in status_clean and "Partially" not in status_clean:
            status_cell.fill = green_fill
        elif "Partially" in status_clean:
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = red_fill

    # Auto-width columns
    for col_num, header in enumerate(headers, 1):
        max_len = max(
            [len(str(header))] + [len(str(r.get(header, ""))) for r in results]
        ) + 4
        detail_sheet.column_dimensions[get_column_letter(col_num)].width = max_len
        # ---------- Bar chart: Match % by Attribute ----------
    bar = BarChart()
    bar.type = "bar"          # horizontal bars, matches your Streamlit chart
    bar.title = "Coverage by Attribute"
    bar.x_axis.title = "Match %"

    last_row = detail_sheet.max_row  # last row of actual data (11 attributes + header)

    cats = Reference(detail_sheet, min_col=1, min_row=2, max_row=last_row)   # Attribute column, skip header
    vals = Reference(detail_sheet, min_col=4, min_row=1, max_row=last_row)   # Match % column, include header for series name

    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 10
    bar.width = 20

    detail_sheet.add_chart(bar, "F2")

    # ---------- Save to memory, not disk ----------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- Streamlit UI ---
uploaded_file = st.file_uploader("Upload BRSR / Annual Report PDF", type="pdf")

if uploaded_file is not None:
    with st.spinner("Scanning PDF for BRSR section..."):
        doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
        hits = find_brsr_section(doc)

    if not hits:
        st.error("Could not locate a BRSR section in this PDF.")
    else:
        section_c_pages = [h["page"] for h in hits if "Section C" in h["keyword"]]

        if not section_c_pages:
            principle_1_pages = []
            for page_num in range(len(doc)):
                text = doc[page_num].get_text()
                if "principle 1:" in text.lower():
                    principle_1_pages.append(page_num)
            section_c_pages = principle_1_pages

        if not section_c_pages:
            st.error("Could not locate the BRSR Principle-wise section in this PDF.")
        else:
            start_page = section_c_pages[-1]
            uploaded_file.seek(0)
            with st.spinner("Extracting and analyzing disclosures..."):
                extracted = extract_text_pdfplumber(uploaded_file, start_page, start_page + 80)
                chunks = chunk_by_principle(extracted)
                full_text = " ".join(chunks.values())
                results = run_gap_analysis(full_text)

            st.success(f"Analysis complete — found {len(chunks)} principle sections.")

            # --- Summary Cards ---
            total_attrs = len(results)
            covered_count = sum(1 for r in results if r["Status"] == "✅ Covered")
            partial_count = sum(1 for r in results if r["Status"] == "🟡 Partially Covered")
            missing_count = sum(1 for r in results if r["Status"] == "🔴 Missing")
            overall_score = round(sum(r["Match %"] for r in results) / total_attrs, 1)

            st.subheader("📋 Summary")
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Overall BRSR Core Score", f"{overall_score}%")
            col2.metric("✅ Covered", covered_count)
            col3.metric("🟡 Partial", partial_count)
            col4.metric("🔴 Missing", missing_count)

            st.divider()

            # --- Charts ---
            gauge_col, donut_col = st.columns([1, 1])

            with gauge_col:
                st.subheader("Overall Score")
                fig_gauge = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=overall_score,
                    number={"suffix": "%"},
                    gauge={
                        "axis": {"range": [0, 100]},
                        "bar": {"color": "#2ecc71"},
                        "steps": [
                            {"range": [0, 50], "color": "#4a1a1a"},
                            {"range": [50, 80], "color": "#4a3a1a"},
                            {"range": [80, 100], "color": "#1a3a1a"}
                        ]
                    }
                ))
                fig_gauge.update_layout(margin=dict(t=30, b=10, l=30, r=30), height=280)
                st.plotly_chart(fig_gauge, use_container_width=True)

            with donut_col:
                st.subheader("Status Breakdown")
                fig_donut = go.Figure(data=[go.Pie(
                    labels=["Covered", "Partial", "Missing"],
                    values=[covered_count, partial_count, missing_count],
                    hole=0.5,
                    marker=dict(colors=["#2ecc71", "#f1c40f", "#e74c3c"])
                )])
                fig_donut.update_layout(
                    showlegend=True,
                    margin=dict(t=10, b=10, l=10, r=10),
                    height=280
                )
                st.plotly_chart(fig_donut, use_container_width=True)

            st.subheader("Coverage by Attribute")
            attr_names = [r["Attribute"] for r in results]
            match_pcts = [r["Match %"] for r in results]
            bar_colors = [
                "#2ecc71" if r["Status"] == "✅ Covered"
                else "#f1c40f" if r["Status"] == "🟡 Partially Covered"
                else "#e74c3c"
                for r in results
            ]

            fig_bar = go.Figure(data=[go.Bar(
                x=match_pcts,
                y=attr_names,
                orientation="h",
                marker=dict(color=bar_colors)
            )])
            fig_bar.update_layout(
                xaxis_title="Match %",
                margin=dict(t=10, b=10, l=10, r=10),
                height=400
            )
            st.plotly_chart(fig_bar, use_container_width=True)

            st.subheader("Attribute Comparison (Radar)")

            benchmark_values = [80] * len(attr_names)  # SEBI reasonable-assurance target line

            fig_radar = go.Figure()

            fig_radar.add_trace(go.Scatterpolar(
                r=match_pcts + [match_pcts[0]],
                theta=attr_names + [attr_names[0]],
                fill='toself',
                name=uploaded_file.name.replace(".pdf", ""),
                line=dict(color="#3498db"),
                fillcolor="rgba(52, 152, 219, 0.4)"
            ))

            fig_radar.add_trace(go.Scatterpolar(
                r=benchmark_values + [benchmark_values[0]],
                theta=attr_names + [attr_names[0]],
                name="Target Benchmark (80%)",
                line=dict(color="#2ecc71", dash="dash"),
                fill=None
            ))

            fig_radar.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.15),
                margin=dict(t=30, b=60, l=60, r=60),
                height=500
            )
            st.plotly_chart(fig_radar, use_container_width=True)
            st.subheader("🌐 Principle → Attribute Breakdown (Sunburst)")

            sunburst_ids = ["BRSR Core"]
            sunburst_labels = ["BRSR Core"]
            sunburst_parents = [""]
            sunburst_values = [0]
            sunburst_colors = ["#1a1f2b"]

            principles = sorted(set(r["Principle"] for r in results))
            for p in principles:
                sunburst_ids.append(p)
                sunburst_labels.append(p)
                sunburst_parents.append("BRSR Core")
                sunburst_values.append(0)
                sunburst_colors.append("#3498db")

            for r in results:
                sunburst_ids.append(r["Attribute"])
                sunburst_labels.append(r["Attribute"])
                sunburst_parents.append(r["Principle"])
                sunburst_values.append(r["Match %"] if r["Match %"] > 0 else 1)
                if r["Status"] == "✅ Covered":
                    sunburst_colors.append("#2ecc71")
                elif r["Status"] == "🟡 Partially Covered":
                    sunburst_colors.append("#f1c40f")
                else:
                    sunburst_colors.append("#e74c3c")

    

            fig_sunburst = go.Figure(go.Sunburst(
                ids=sunburst_ids,
                labels=sunburst_labels,
                parents=sunburst_parents,
                values=sunburst_values,
                branchvalues="remainder",
                marker=dict(colors=sunburst_colors),
                hovertemplate='<b>%{label}</b><br>Match: %{value}%<extra></extra>'
            ))
            fig_sunburst.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=550)
            st.plotly_chart(fig_sunburst, use_container_width=True)
            st.subheader("🔲 Quick Status Grid")

            import math
            cols_per_row = 4
            rows_needed = math.ceil(len(results) / cols_per_row)

            status_color_map = {
                "✅ Covered": "#2ecc71",
                "🟡 Partially Covered": "#f1c40f",
                "🔴 Missing": "#e74c3c"
            }

            grid_html = "<div style='display:grid; grid-template-columns: repeat(4, 1fr); gap: 10px;'>"
            for r in results:
                color = status_color_map[r["Status"]]
                tile = (
                    f"<div style='background-color:{color}; border-radius:10px; padding:14px; text-align:center;'>"
                    f"<div style='font-size:13px; font-weight:600; color:#0d1117;'>{r['Attribute']}</div>"
                    f"<div style='font-size:20px; font-weight:800; color:#0d1117;'>{r['Match %']}%</div>"
                    f"</div>"
                )
                grid_html += tile
            grid_html += "</div>"

            st.markdown(grid_html, unsafe_allow_html=True)

            st.divider()

            st.subheader("📄 Detailed Attribute Coverage")

            all_principles = ["All Principles"] + sorted(set(r["Principle"] for r in results))
            selected_principle = st.selectbox("🔍 Drill down by Principle", all_principles)

            if selected_principle == "All Principles":
                filtered_results = results
            else:
                filtered_results = [r for r in results if r["Principle"] == selected_principle]

            table_view = [
                {
                    "Attribute": r["Attribute"],
                    "Principle": r["Principle"],
                    "Status": r["Status"],
                    "Match %": r["Match %"]
                }
                for r in filtered_results
            ]
            st.table(table_view)

            st.markdown("**🔎 Evidence — the actual sentence behind each match**")
            st.caption(
                "Keywords found in the filing are shown with the sentence they came from. "
                "If a keyword appears only in a negated context (e.g. \"no water recycling\"), "
                "it's counted as a gap, not a match — shown here so you can verify it yourself."
            )
            for r in filtered_results:
                if not r["Evidence"]:
                    continue
                with st.expander(f"{r['Status']} · {r['Attribute']} — evidence"):
                    for kw in r["Matched Keywords"]:
                        sentence = r["Evidence"].get(kw)
                        if sentence:
                            st.markdown(f"✅ **`{kw}`** — _{sentence}_")
                    for kw in r["Missing Keywords"]:
                        sentence = r["Evidence"].get(kw)
                        if sentence:
                            st.markdown(f"🚫 **`{kw}`** — found only in negated context — _{sentence}_")

            if selected_principle != "All Principles":
                sel_scores = [r["Match %"] for r in filtered_results]
                sel_avg = round(sum(sel_scores) / len(sel_scores), 1)
                st.info(f"**{selected_principle}** average coverage: **{sel_avg}%** across {len(filtered_results)} attributes")
            st.divider()

            # --- AI-Generated Compliance Gap Narratives ---
            st.subheader("🤖 AI Compliance Gap Narratives")
            st.caption(
                "For each Partial or Missing attribute, generate a specific auditor-style "
                "narrative: what's missing, why SEBI cares, and what to disclose next cycle."
            )

            gap_results = [r for r in results if r["Status"] != "✅ Covered"]

            if not gap_results:
                st.success("No gaps detected — every BRSR Core attribute is covered. Nothing to narrate.")
            else:
                narrative_key = f"narratives_{uploaded_file.name}"
                if narrative_key not in st.session_state:
                    st.session_state[narrative_key] = {}

                if st.button(f"✨ Generate narratives for {len(gap_results)} gap(s)"):
                    client = get_anthropic_client()
                    if client is None:
                        st.warning(
                            "Add your Anthropic API key in the sidebar under 'AI Narratives' to use this feature."
                        )
                    else:
                        progress = st.progress(0.0, text="Generating narratives...")
                        for i, r in enumerate(gap_results):
                            narrative = generate_gap_narrative(
                                client,
                                attribute=r["Attribute"],
                                principle=r["Principle"],
                                status=r["Status"].replace("🟡 ", "").replace("🔴 ", ""),
                                match_pct=r["Match %"],
                                matched_keywords=r["Matched Keywords"],
                                missing_keywords=r["Missing Keywords"],
                                company_name=uploaded_file.name.replace(".pdf", "")
                            )
                            st.session_state[narrative_key][r["Attribute"]] = narrative
                            progress.progress((i + 1) / len(gap_results), text=f"Generated {i + 1}/{len(gap_results)}")
                        progress.empty()

                if st.session_state[narrative_key]:
                    for r in gap_results:
                        narrative = st.session_state[narrative_key].get(r["Attribute"])
                        if narrative:
                            with st.expander(f"{r['Status']} · {r['Attribute']} ({r['Principle']}) — {r['Match %']}%"):
                                st.write(narrative)

            st.divider()

            # --- Downloadable PDF Report ---
            narrative_key = f"narratives_{uploaded_file.name}"
            pdf_buffer = generate_pdf_report(
                results, overall_score, covered_count, partial_count, missing_count,
                company_name=uploaded_file.name.replace(".pdf", ""),
                narratives=st.session_state.get(narrative_key)
            )
            excel_buffer = generate_excel_report(
                results, overall_score, covered_count, partial_count, missing_count,
                company_name=uploaded_file.name.replace(".pdf", "")
            )

            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    label="📥 Download PDF Report",
                    data=pdf_buffer,
                    file_name=f"BRSR_Gap_Analysis_{uploaded_file.name.replace('.pdf', '')}.pdf",
                    mime="application/pdf"
                )
            with dl_col2:
                st.download_button(
                    label="📊 Download Excel Report",
                    data=excel_buffer,
                    file_name=f"BRSR_Gap_Analysis_{uploaded_file.name.replace('.pdf', '')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )